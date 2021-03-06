from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import Workbook, load_workbook
import os
from werkzeug.utils import secure_filename
app = Flask(__name__)

UPLOAD_FOLDER = './static/test_pic/'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['SECRET_KEY'] = 'thisissecret'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

class temporary:
    # table buat nama tabel yang di buat
    # columns buat nyimpen angka(type data belum integer) yang diinput sama user
    # store buat nyimpen semua data di setiap row pada setiap column. contoh : user input 2 column brarti di dlm store =(row1column1,row1column2,row2column1,row2column2,row3column1,row3column2,dst sampe usernya puas)
    table = None
    columns = None
    store = []
    changer = []
    added = None
    lister = []
    viewing = []
    changes = 0
    check = []


    # coba panggil function
    def add_function(A):
        result = sum(A)
        return result
    # syntax manggil def add_function di html
    app.jinja_env.globals.update(add_function=add_function)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class Tbs:
    def __init__(tb_att, column):
        tb_att.column = column

@app.route('/')
def home():

    return render_template('index1.html')
@app.route('/home_A')
def home_A():

    return render_template('index.html')


# Route for handling the login page logic
@app.route('/login', methods=['POST'])
def do_admin_login():
    if request.form['password'] == '0000' and request.form['username'] == 'root':
        return render_template('index.html')
    else:
        flash('wrong password!')
        return home()


@app.route('/process', methods=['get'])
def process():
    selected = 1
    # just to make sure everything restart
    temporary.table = None
    temporary.columns = None
    temporary.store = []
    temporary.changer = []
    temporary.added = None
    return render_template('index.html', selected=selected)
@app.route('/process2', methods=['POST','GET'])
def process2():
    try:
        Tb_name = request.form['filename']
        temporary.check = Tb_name
        columns = request.form['columns']
        columns = int(columns)
    except:
        #untuk nyelesain masalah undo saat awal create table alurnya dari app route /back
        Tb_name = temporary.check

    try:
        wb = load_workbook(filename='data.xlsx')
    except:
        wb = Workbook()
        wb.save(filename='data.xlsx')

    for x in wb.sheetnames:
        if Tb_name == x:
            warning = 'this table is exist please delete or use other name'
            return render_template('index.html', warning=warning)

    # temporary adalah class untuk nyimpen data sementara bisa dilihat awal diatas
    try:
        temporary.table = Tb_name
        temporary.columns = columns
    except:
        pass
    return render_template('single.html', Tb_name=temporary.table, columns=temporary.columns)
@app.route('/storing_data', methods=['POST'])
def storing_data():
    Tb_name = temporary.table
    columns = temporary.columns
    columns = int(columns)
    for x in range(columns):
        # x perlu dijadiin string karena request.form perlunya string bkan integer
        x = str(x)
        data = request.form[x]
        temporary.store.append(data)
    stored = temporary.store
    #changer dipakai untuk memangil value setiap row, penjelasan lebih lanjut di single.html
    if temporary.added == None:
        temporary.changer.append(-columns)
        temporary.added = "yes"
        #kurang 1 karena isi changer awal diisi dgn -column
        for x in range(columns - 1):
            temporary.changer.append(columns)
    else:
        for x in range(columns):
            temporary.changer.append(columns)
    changer = temporary.changer

    #rows adalah jumlah row yang ada, pemakaian ada di single.html
    try:
        rows = len(temporary.store)/columns
        rows = int(rows)
    except:
        rows = int(0)
    # ini for buat mangil input dari html ny kalo diliat di html <input name=x> x itu range(column),
    # contohnya user input column 3 brarti setiap row dia masukin 3 data yaitu row1column1,row1column2,row1column3>>>nah x nya itu jadi x=1,2,3 >> 1 buat manggil row1column1, 2 buat manggil row1column2, dst


    return render_template('single.html', Tb_name=Tb_name, columns=columns, changer=changer, rows=rows, stored=stored)

@app.route('/save_reset')
def save_reset():
    # if buat mastiin user masuk ke single.html dari dpan bukan dari blakang karena di browser ada back button setelah user kluar dari single.html temporary semua value di dlmnya bakal ke reset jadi bisa ketangkep sama if kalo dy back button
    if temporary.table == None:
        warning2 = "table name not detected please create new table or edit table"
        return render_template('index.html', warning2=warning2)
    else:
        # pertama2 load dulu excel namanya data.xlsx di try kalo ternyata excelnya blom ke buat masuk ke except dy otomatis buat excel yang namanya data.xlsx
        try:
            wb = load_workbook(filename='data.xlsx')
        except:
            wb = Workbook()
            wb.save(filename='data.xlsx')
        #perlu Tb_name karena rencana nya pake 1 excel doank dimana didlm 1 excel itu ada banyak page/sheet, sheet nya itu diksh nama tabel yang diinput, jadi kalo mw hapus tinggal hapus nama sheet nya aja
        Tb_name = temporary.table
        ws = wb.create_sheet(Tb_name)
        #ini algorithm ngemisahin data di temporary.store
        columns = temporary.columns
        columns = int(columns)
        print('temporary debug in save_reset=')
        print(temporary.store)
        # row quantitiy untuk mengetahui jumlah row. dipake untuk openpyxl masukin value ke specific cell
        if temporary.store == []:
            warning4 = 1
            return render_template("index.html", warning4 = warning4)
        row_quantity = len(temporary.store)/columns
        row_quantity = int(row_quantity)
        for x in range(columns):
            # +1 gara x awal itu 0 bukan 1
            index = x + 1
            #ngeloop sebanyak x yaitu sebanyak kolom yang di pilih
            ws.insert_cols(0)
            for y in range(row_quantity):
                #ngeloop terus sebanyak row yang ada

                selector = str(y + 1)

                # dibawah ini ws['A' + selector], jika selectornya 1 jadi ws['A1'] artinya cell A1 selector didapat dari row_quantity
                ws['A' + selector] = temporary.store[-index]
                print('A'+selector)
                # untuk indexing kita dari belakang temporary.store[-index] ada minus yang artinya ambil dari paling blakang mengapa?
                # karena kita menggunakan ws.insert_rows(0) yang artinya buat kolom baru di existing kolom ke 0 artinya kolom pertama yang lama geser jadi kolom kedua dan kolom pertama jadi yang baru saja dibuat
                # makanya kita indexing dri blakang gara2 geser yang dimasukin awal jadi yang terakhir.
                index = index + columns
                #index diata diupdate dengan rumus yg terter diatas karena dari row1kolom1 ke row2kolom1 harus lompat sebanyak (column-1)
        # untuk jadiin row trakir ke row pertama dst ato g kebalik
        again = []
        for idr, row in enumerate(ws.rows):
            reverse = []
            idx2 = idr + 1
            for cols in ws.columns:
                reverse.append(cols[-idx2].value)
            again.append(reverse)
        ws = wb.get_sheet_by_name(temporary.table)
        wb.remove_sheet(ws)
        ws = wb.create_sheet(temporary.table)
        for item in again:
            ws.append(item)
        wb.save(filename='data.xlsx')
        # reset smua value
        temporary.table = None
        temporary.columns = None
        temporary.store = []
        temporary.changer = []
        temporary.added = None

        return redirect(url_for('home_A'))
@app.route('/cancel')
def cancel():
    temporary.store = []
    temporary.table = None
    temporary.columns = None
    temporary.changer = []
    temporary.added = None
    temporary.lister = []
    temporary.viewing = []
    temporary.changes = 0
    if 'viewed' in session:
        session.pop('viewed', None)  # delete visits
    return redirect(url_for('home_A'))
@app.route('/back')
def back():
    Tb_name = temporary.table
    columns = temporary.columns
    columns = int(columns)
    print(temporary.store)
    for x in range(columns):
        try:
            #hapus row trakir cnth user input 6 kolom brarti 1 row ada 6 value di pop 6 kali biar 6 value trakir di temporary store ilang
            temporary.store.pop()
            temporary.changer.pop()
        except:
            return redirect(url_for('process2'))
    stored = temporary.store
    changer = temporary.changer
    #rows adalah jumlah row yang ada, pemakaian ada di single.html
    try:
        rows = len(temporary.store)/columns
        rows = int(rows)
    except:
        rows = int(0)
    return render_template('single.html', Tb_name=Tb_name, columns=columns, changer=changer, rows=rows, stored=stored)
@app.route('/process_delete')
def process_delete():
    try:
        wb = load_workbook(filename='data.xlsx')
        Q = len(wb.sheetnames)
        Q = int(Q)
        temporary.lister = []
        lister = []
        # perlu mulai dari satu soalnya sheet paling awal di excelny jgn di apus ato g nanti [1:1] ga bisa baca sheet ny kluar ny NONE
        for x in wb.sheetnames[1:Q]:
            lister.append(x)
            temporary.lister.append(x)
    except:
        wb = "there is nothing to delete"
        lister = wb

    return render_template('index.html', lister=lister)
@app.route('/deleted', methods=['POST'])
def deleted():
    lister = temporary.lister
    wb = load_workbook(filename='data.xlsx')
    # perlu mulai dari satu soalnya sheet paling awal di excelny jgn di apus
    if request.method == 'POST':
        print(lister)
        for x in lister:
            x = str(x)
            try:
                if request.form['submit_button'] == x:
                    for item in wb.sheetnames:

                        if item == x:

                            item = str(item)
                            temp = wb[item]
                            wb.remove_sheet(temp)
                            #?????????????????????????????????
                            security = []
                            for filename in os.listdir("./static/test_pic"):
                                if filename == (item + ".png"):
                                    security.append(filename)
                                    os.remove("./static/test_pic/" + item + ".png")
                            #????????????????????????????????
                            else:
                                lister.remove(x)
                                wb.save(filename='data.xlsx')
                                return render_template('index.html', lister=lister)
            except:
                print("excep")
                pass  # do something else
    return render_template('index.html', lister=lister)
@app.route('/view_table')
def view_table():
    # process awal nampilin semua table
    try:
        wb = load_workbook(filename='data.xlsx')
        #untuk memastikan temporary.lister kosong
        temporary.lister = []
        #isi lister
        lister1 = []
        Q = len(wb.sheetnames)
        Q = int(Q)
        # perlu mulai dari satu soalnya sheet paling awal di excelny jgn di apus
        for x in wb.sheetnames[1:Q]:
            lister1.append(x)
            temporary.lister.append(x)
        return render_template('index.html', lister1=lister1)
    except:
        warning2 = "table name not detected please create new table or edit table"
        return render_template('index.html', warning2=warning2)
@app.route('/view_process', methods=['POST'])
def view_process():
    #proces tengah untuk dapetin nama table yang nanti akan di panggil untuk root selanjutnya
    lister = temporary.lister
    wb = load_workbook(filename='data.xlsx')
    temporary.table = None
    if request.method == 'POST':
        for x in lister:
            x = str(x)
            try:
                if request.form['submit_button'] == x:
                    for item in wb.sheetnames:

                        if item == x:
                            Tb_name = str(item)
                            temporary.table = Tb_name
                            ws = wb[Tb_name]
                            # make sure temporary.viewing is empty
                            temporary.viewing = []
                            temporary.columns = len(ws[1])
                            for idx, rows in enumerate(ws.rows):
                                # isi dri temporary.viewing adalah ([row 1],[row 2], ....) dimana row 1 isi nya [col1,col2,col3,...]>>place_temp
                                place_temp = []
                                for cols in ws.columns:
                                    place_temp.append(cols[idx].value)
                                temporary.viewing.append(place_temp)

                            return redirect(url_for('viewed'))
            except:
                pass  # do something else
    return '<h1>view_process error</h1>'
@app.route('/viewed')
def viewed():
    if 'viewed' in session:
        session['viewed'] = session.get('viewed') + 1  # reading and updating session data
    else:
        session['viewed'] = 1  # setting session data
        # buat ngecek apakah user menggunakan backk button browser setiap sesi bertambah maka seharusnya changes bertambah
        # param changes harus =1 karena session['viewed'] awal ny 1 changes ny msh 0

    param_changes = session['viewed'] - temporary.changes

    if param_changes == 1:
        #page tabel yang terpilih temporary.table dan temporary.viewing sdh terisi sebelum masuk route ini.

        Tb_name = temporary.table
        fnally = temporary.viewing
        print('fnally={}'.format(fnally))
        index = []
        for idx, item in enumerate(fnally):
            index.append(idx)
        # combine = [[1,[row1],[2,[row2],....]]] biar bisa dpt idx ny di html soalnya jinja g bisa pake enumerate
        combine = zip(index, fnally)
        # untuk <input> ny sesuai sebanyak kolom

        columns = temporary.columns
        return render_template('styles.html', fnally=combine, Tb_name=Tb_name, columns=columns)
    else:
        warning3 = 'please try again'
        session.pop('viewed', None)  # delete visits
        temporary.changes = 0

        return render_template('index.html', warning3=warning3)
@app.route('/edit_to_del',methods=['POST'])
def edit_to_del():
    # process layer terakir untuk penghapusan sebelm user submit dan di save excelnya
    fnally = temporary.viewing
    try:
        # bwt dapetin index row yg mw dihapus
        for idx, x in enumerate(fnally):
            stridx = str(idx)

            if request.form['submit_button'] == stridx:
                temporary.viewing.pop(idx)
                temporary.changes = temporary.changes + 1
                return redirect(url_for('viewed'))
    except:
        pass  # do something else
    return "<h1>error in edit_to_del</h1>"
@app.route('/edit_to_insert',methods={'POST'})
def edit_to_insert():
    place_temp = []
    #temporary.viewing untuk ambil column
    for idx, x in enumerate(temporary.viewing[0]):
        idx = str(idx)
        send = request.form[idx]
        place_temp.append(send)
    temporary.viewing.append(place_temp)
    print(temporary.viewing)
    temporary.changes = temporary.changes + 1
    return redirect(url_for('viewed'))
@app.route('/saving')
def saving():
    if temporary.table == None:
        warning3 = 'please try again'

        return render_template('index.html', warning3=warning3)
    else:
        wb = load_workbook(filename='data.xlsx')
        # hapus sheet buat baru dgn data baru dari temporary.viewing
        std = wb.get_sheet_by_name(temporary.table)
        if temporary.viewing == []:
            wb.remove_sheet(std)

            session.pop('viewed', None)  # delete visits
            temporary.viewing = []
            temporary.columns = None
            temporary.changes = 0
            temporary.table = None
            wb.save('data.xlsx')
            return render_template('index.html')
        else:
            wb.remove_sheet(std)

            print(temporary.viewing)
            wb.save('data.xlsx')
            ws = wb.create_sheet(temporary.table)
            for row in temporary.viewing:
                ws.append(row)
                print('debug={}'.format(row))
            wb.save('data.xlsx')
            session.pop('viewed', None)  # delete visits
            temporary.viewing = []
            temporary.columns = None
            temporary.changes = 0
            temporary.table = None
            return render_template('index.html')
@app.route('/blog')
def blog():
    return render_template('blog.html')
@app.route('/tabloid')
def tabloid():
    return render_template('single.html')
@app.route('/styles')
def style():
    return render_template('styles.html')

@app.route('/gallery')
def gallery():
    folder = './static/test_pic'
    images = []
    for filename in os.listdir(folder):
        images.append(filename)
    print(images)
    return render_template('gallery1.html', images=images)
@app.route('/Display_A')
def Display_A():
        try:
            wb = load_workbook(filename='data.xlsx')
            # persiapan untuk buat zip bentuknya ntr jadi = (1,sheet1),(2,sheet2)
            sheet = []
            index = []
            count = 0
        except:
            return 'no workbook found!'
        try:
            for idx, i in enumerate(wb.sheetnames):
                count = count + 1
                sheet.append(i)
                index.append(idx)
            # ilangin sheet awal karena g dipake
            sheet.pop(0)
            index.pop(0)
            # buat zip untuk enumerate di html nanti
            sheets = zip(index, sheet)
            # all_column_class=[class1,class2] class ny diatas Tbs
            all_column_class = []
            for page in wb.sheetnames:
                activate = wb[page]
                column_holding = []

                for x in activate.rows:
                    column_val = []
                    # x nya musti di for lg karna x itu 1row[col1,col2,..] jadi list bukan cell
                    for val in x:
                        column_val.append(val.value)
                    column_holding.append(column_val)
                    print("debug2={}".format(x))
                ##print("debug={}".format(column_holding))
                #fix bug ngambil column hrsny ambil row di column_holding


                all_column_class.append({page: Tbs(column_holding)})
            # load all picture
            pictures = []
        except:
            return 'failed in building class'
        for filename in os.listdir("./static/test_pic"):
            pictures.append(filename)
            count = count + 1

        return render_template('Display_A.html', all_column_class=all_column_class, sheets=sheets, pictures=pictures, count=count)

@app.route('/save_pic', methods=['GET', 'POST'])
def save_pic():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return redirect(url_for('save_pic',
                                    filename=filename))
    flash('file uploaded')
    return render_template('index.html')

@app.route('/delete_pic')
def delete_pic():
    # delete picture
    images_holder = []
    for image in os.listdir("./static/test_pic/"):
        images_holder.append(image)
    print(images_holder)

    return render_template('index.html', image_holder=images_holder)

@app.route('/delete_pic_process', methods=['POST'])
def delete_pic_process():
    # delete picture process then loop back to delete_pic route
    images_holder = []
    for image in os.listdir("./static/test_pic/"):
        images_holder.append(image)
    for trgt in images_holder:
        if trgt in request.form:
            # deleting target picture
            os.remove("./static/test_pic/" + trgt)
            return redirect(url_for('delete_pic'))
    return

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5005, debug=True)
